## Libs
import openpyxl
from openpyxl import load_workbook
import sys
import os
import re
import datetime as dt
import streamlit as st
import tkinter as tk
from tkinter import filedialog

st.title("Gerador de planilha LC a partir do Mapa Caixa")
## Get folder and mapacaixa files

# select_mapacaixa_folder_button = st.button("Selecione a pasta com os arquivos do Mapa Caixa")
st.session_state.folder = st.text_input(
    "Copie e Cole aqui o caminho da pasta que contem todos os mapa-caixa"
)

if st.session_state.folder:
    # st.session_state.app = QApplication(sys.argv)
    # st.session_state.folder = QFileDialog.getExistingDirectory(None, "Select Folder")
    # root = tk.Tk()
    # root.withdraw()  # Hide the Tkinter root window
    print(st.session_state.folder)
    st.write(st.session_state.folder)
    # st.session_state.folder = os.path.join(st.session_state.folder.name.split("/")[:-1])
    # print(st.session_state.folder)
    # root.destroy()
    st.session_state.files = os.listdir(st.session_state.folder)
    print(st.session_state.files)
    st.session_state.sheetnames = [
        f for f in st.session_state.files if re.match(r"[mM].+?\.xlsx$", f)
    ]
    st.session_state.night_sheetnames = [
        sheet
        for sheet in st.session_state.sheetnames
        if re.search("[nN][oO][iI][tT][eE]", sheet)
    ]
    st.session_state.day_sheetnames = [
        sheet
        for sheet in st.session_state.sheetnames
        if sheet not in st.session_state.night_sheetnames
    ]
    st.write("Arquivos de dia:", st.session_state.day_sheetnames)
    st.write("Arquivos de noite:", st.session_state.night_sheetnames)

start_copy_cell = "AE7"
end_copy_cell = "AE60"

## Create new LC
now = dt.datetime.now()
month = st.number_input("Mês", min_value=1, max_value=12, value=now.month)
year = st.number_input("Ano", min_value=2024, max_value=9999, value=now.year)
first_dt_cell = "C4"
first_dt_weekname_cell = "C3"
new_lc_namefile = f"LC_{month}_{year}.xlsx"

## Open LC Sheets
# select_lc_button = st.button("Selecione a planilha base do LC")
st.session_state.base_lc_filename = st.file_uploader(
    "Escolha a planilha base do LC", type=["xlsx", "csv", "txt"]
)
if st.session_state.base_lc_filename:
    # st.session_state.base_lc_filename = QFileDialog.getOpenFileName(None, "Select Base LC File", filter="Excel Files (*.xlsx)")[0]
    # root = tk.Tk()
    # root.withdraw()  # Hide the Tkinter root window
    # st.session_state.base_lc_filename = filedialog.askopenfilename(title="Selecionar planilha base do LC",
    #                                                                filetypes=[("Excel files", "*.xlsx")])
    st.session_state.new_lc = load_workbook(st.session_state.base_lc_filename)
    st.session_state.lc_sheetnames = st.session_state.new_lc.sheetnames
    st.session_state.lc_day_sheetname = [
        sheet for sheet in st.session_state.lc_sheetnames if "dia" in sheet.lower()
    ][0]
    st.session_state.lc_day_sheet = st.session_state.new_lc[
        st.session_state.lc_day_sheetname
    ]
    st.session_state.lc_night_sheetname = [
        sheet for sheet in st.session_state.lc_sheetnames if "noite" in sheet.lower()
    ][0]
    st.session_state.lc_night_sheet = st.session_state.new_lc[
        st.session_state.lc_night_sheetname
    ]
    ## Set the dates correctly
    first_dt = dt.datetime(day=1, month=month, year=year)
    first_dt_weekname = first_dt.strftime("%A")[0:3]
    st.session_state.lc_day_sheet[first_dt_cell] = first_dt
    st.session_state.lc_day_sheet[first_dt_weekname_cell] = first_dt_weekname


## Copy and paste values from MapaCaixa to LC


def copy_and_paste_values(lc_sheet, sheetnames, start_copy_cell, end_copy_cell):
    for sheetname in sheetnames:
        filename = os.path.join(st.session_state.folder, sheetname)
        workbook = load_workbook(filename, keep_vba=True, data_only=True)
        active_sheet = workbook[workbook.sheetnames[0]]
        # Extract values from the range
        values = [cell[0].value for cell in active_sheet[start_copy_cell:end_copy_cell]]
        sheet_dt = re.findall(r"\d+", sheetname)
        sheet_day = int(sheet_dt[0])
        # sheet_dt = dt.datetime(day=sheet_dt[0], month=sheet_dt[1], year=sheet_dt[2])
        paste_column = openpyxl.utils.get_column_letter(
            2 + sheet_day
        )  # Starts on C and C is the 3rd column
        start_paste_cell = f"{paste_column}8"
        end_paste_cell = f"{paste_column}61"
        print(start_paste_cell, end_paste_cell)
        print(values)
        # Copy values to the new sheet
        cells = lc_sheet[start_paste_cell:end_paste_cell]
        for i, cell in enumerate(cells):
            cell[0].value = values[i]
        workbook.close()


create_lc_button = st.button(f"Criar LC para o mês {month} de {year}")
if create_lc_button:
    with st.spinner("Criando LC..."):
        copy_and_paste_values(
            st.session_state.lc_day_sheet,
            st.session_state.day_sheetnames,
            start_copy_cell,
            end_copy_cell,
        )
        copy_and_paste_values(
            st.session_state.lc_night_sheet,
            st.session_state.night_sheetnames,
            start_copy_cell,
            end_copy_cell,
        )

        ## Save new LC
        st.session_state.new_lc.save(new_lc_namefile)
        st.session_state.new_lc.close()
        st.success("LC criado com sucesso!")
